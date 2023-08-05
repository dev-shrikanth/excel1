import requests
import openpyxl
from openpyxl.workbook.defined_name import DefinedName


def get_api_response(api_url):
    response = requests.get(api_url)
    if response.status_code == 200:
        return response.json()
    else:
        print(
            f"Failed to get API response for URL: {api_url}. Status Code: {response.status_code}"
        )
        return None


def write_json_list_to_worksheet(worksheet, json_list, start_row):
    # Write headers in the worksheet
    headers = list(json_list[0].keys())
    for col_num, header in enumerate(headers, start=1):
        worksheet.cell(row=start_row, column=col_num, value=header)

    # Write data to the worksheet
    for row_num, item in enumerate(json_list, start=start_row + 1):
        for col_num, header in enumerate(headers, start=1):
            worksheet.cell(row=row_num, column=col_num, value=item.get(header, ""))

    # Calculate the end row of the data range
    end_row = start_row + len(json_list)

    return end_row  # Return the end row


if __name__ == "__main__":
    workbook = openpyxl.Workbook()
    backend_sheet = workbook.active
    backend_sheet.title = "backend"

    # Accept 5 API URLs and 5 named ranges
    api_urls = []
    named_ranges = []
    start_row = 1  # Set initial start row to 1

    for i in range(5):
        api_url = input(f"Enter API URL {i + 1}: ")
        api_urls.append(api_url)
        named_range = input(f"Enter named range for API {i + 1}: ")
        named_ranges.append(named_range)

        json_response = get_api_response(api_url)

        if json_response and isinstance(json_response, list):
            write_json_list_to_worksheet(backend_sheet, json_response, start_row)
            # Create a named range
            named_range_ref = f"{backend_sheet.title}!$A${start_row}:${openpyxl.utils.get_column_letter(len(json_response[0]))}${start_row+len(json_response)-1}"
            named_range_obj = DefinedName(name=named_range, attr_text=named_range_ref)
            workbook.defined_names.add(named_range_obj)
            print(f"Named range '{named_range}' created for the data from API {i + 1}.")
            start_row += (
                len(json_response) + 2
            )  # Update start_row for the next API call
        else:
            print(
                f"Invalid API response for URL {api_url}. The API should return a list of dictionaries."
            )

    # Save the Excel file
    output_filename = "output.xlsx"
    workbook.save(output_filename)
    print(f"\nJSON data successfully written to {output_filename}")
    print("Named ranges created for the data from multiple API calls.")