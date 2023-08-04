import requests
import openpyxl
from openpyxl import load_workbook


def get_api_response(api_url):
    response = requests.get(api_url)
    if response.status_code == 200:
        return response.json()
    else:
        print(f"Failed to get API response. Status Code: {response.status_code}")
        return None


def write_json_list_to_excel(json_list, output_filename):
    workbook = openpyxl.Workbook()
    sheet = workbook.create_sheet("Blogs")
    sheet.title = "Blogs List"

    # Write headers in the worksheet
    headers = list(json_list[0].keys())
    for col_num, header in enumerate(headers, start=1):
        sheet.cell(row=1, column=col_num, value=header)

    # Write data to the worksheet
    for row_num, item in enumerate(json_list, start=2):
        for col_num, header in enumerate(headers, start=1):
            sheet.cell(row=row_num, column=col_num, value=item.get(header, ""))

    # Save the Excel file
    workbook.save(output_filename)
    print(f"JSON data successfully written to {output_filename}")


if __name__ == "__main__":
    api_url = input("Enter the API URL: ")
    json_response = get_api_response(api_url)

    if json_response and isinstance(json_response, list):
        output_filename = input("Enter the output Excel filename (e.g., output.xlsx): ")
        write_json_list_to_excel(json_response, output_filename)
    else:
        print("Invalid API response. The API should return a list of dictionaries.")
